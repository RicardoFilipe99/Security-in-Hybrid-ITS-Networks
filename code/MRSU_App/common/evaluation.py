import logging
import timeit

from enum import Enum
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


class EvaluationEnum(Enum):
    NONE = 0
    TotalComputationTime = 1
    SecurityComputationTime = 2
    AvgTimeProcessingIncreasingStationsDensity = 3
    NetworkDelayReceiver = 4
    TransmissionToMobileNetworkDelay = 5
    TransmissionToG5NetworkDelay = 6
    TransmissionToXferViaWSS = 7

    def __str__(self):
        return self.name


class Evaluator:
    """
    Evaluator object that allows to evaluate the application performance
    """

    def __init__(self, evaluation_mode):
        self.evaluation_mode = evaluation_mode
        self.network_delay_transmitter_start = None
        self.network_delay_transmitter_end = None

    def execution_performance(self, execution_lambda, identifier):
        """
        Executes a lambda, calculating the processing time and saving it in an Excel file.

        @param execution_lambda: Lambda function to be executed
        @param identifier: Identifies the Excel column name to save the performance result
        """
        loop = 1
        result = timeit.timeit(execution_lambda, number=loop)

        average_time_seconds = result / loop
        average_time_ms = average_time_seconds * 1000  # Convert to milliseconds

        logging.debug(f'Average processing time: {average_time_ms:.6f} ms')

        if self.evaluation_mode == EvaluationEnum.TotalComputationTime:
            file_to_save_results = "rsu_middleware_total_computation_time.xlsx"
        elif self.evaluation_mode == EvaluationEnum.SecurityComputationTime:
            file_to_save_results = "rsu_middleware_security_computation_time.xlsx"
        else:
            raise Exception(f"Unknown evaluation mode: {self.evaluation_mode}")

        self.write_to_excel(
            filename=file_to_save_results,
            col_name=identifier,
            value=average_time_ms
        )

    def write_to_excel(self, filename, col_name, value):
        """
        Writes a value to an Excel file.

        @param filename: Excel filename
        @param col_name: Excel column
        @param value: Value to write
        """
        try:
            workbook = load_workbook(filename=filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active

        # Find the specified column's index
        col_index = None
        for cell in sheet[1]:
            if cell.value == col_name:
                col_index = cell.column

        # If the specified column does not exist, create it after the last one
        if col_index is None:
            if sheet.max_column == 1:
                for cell in sheet[1]:
                    if cell.value is None:
                        col_index = 1
                        break
                    col_index = sheet.max_column + 1
                    break
            else:
                col_index = sheet.max_column + 1
            sheet.cell(row=1, column=col_index, value=col_name)

        # Find the column's next row
        count = 0
        for row in sheet.iter_cols(min_row=1, min_col=col_index, max_col=col_index):
            for cell in row:
                if cell.value:
                    count += 1

        next_available_row = count + 1

        # Write the value and save
        sheet.cell(row=next_available_row, column=col_index, value=value)
        workbook.save(filename=filename)

    def save_transmission_delay(self, sec_prot_designation):
        """
        Save a transmission latency measurement into an Excel file.

        @param sec_prot_designation: Security approach designation
        """
        filename = "rsu_middleware_transmission_delay.xlsx"

        # Calculate total time
        difference_in_ns = (self.network_delay_transmitter_end - self.network_delay_transmitter_start) / 2
        milliseconds_difference = difference_in_ns / 1e6

        self.write_to_excel(filename=filename, col_name=f"{self.evaluation_mode}_{sec_prot_designation}",
                            value=milliseconds_difference)
